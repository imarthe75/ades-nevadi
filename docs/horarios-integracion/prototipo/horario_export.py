# -*- coding: utf-8 -*-
"""Construye el Excel del horario a partir de asignacion.json y verifica reglas."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DAYS = ['Lun','Mar','Mié','Jue','Vie']
GROUPS = ['1A','1B','2A','2B','3A','3B','4A','4B','5A','5B','6A','6B']
def grade(g): return int(g[0])
def turno(g): return 1 if grade(g)<=3 else 2

curri = {
 1: {'Lecto':7,'Español':1,'Matemáticas':7,'Inglés':4,'Socioemocional':1,'Desarrollo Comunitario':1,'Educación Física':2,'Computación':1,'Artes':1,'Formación':1,'Conocimiento':1,'Proyectos':2},
 2: {'Lecto':5,'Español':3,'Matemáticas':7,'Inglés':4,'Socioemocional':1,'Desarrollo Comunitario':1,'Educación Física':2,'Computación':1,'Artes':1,'Formación':1,'Conocimiento':1,'Proyectos':2},
 3: {'Español':6,'Matemáticas':7,'Inglés':4,'Socioemocional':1,'Desarrollo Comunitario':1,'Educación Física':2,'Computación':1,'Artes':1,'Formación':1,'Entidad':1,'Conocimiento':2,'Proyectos':2},
 4: {'Español':5,'Matemáticas':6,'Inglés':4,'Socioemocional':1,'Desarrollo Comunitario':1,'Educación Física':2,'Computación':1,'Artes':1,'Formación':1,'Conocimiento':2,'Proyectos':3,'Historia':1,'Geografía':1},
}
curri[5]=dict(curri[4]); curri[6]=dict(curri[4])

def blocks_for(t, day):
    if day=='Vie': return [('B1',480),('B2',540),('B3',630),('B4',690),('B5',750)]
    if t==1: return [('B1',480),('B2',540),('B3',630),('B4',690),('B5',810),('B6',870)]
    return [('B1',480),('B2',540),('B3',630),('B4',690),('B5',750),('B6',870)]

START_TICKS = {480:(0,1),540:(2,3),630:(5,6),690:(7,8),750:(9,10),810:(11,12),870:(13,14)}
ROW_LABELS = ['8:00–8:30','8:30–9:00','9:00–9:30','9:30–10:00','10:00–10:30',
              '10:30–11:00','11:00–11:30','11:30–12:00','12:00–12:30','12:30–1:00',
              '1:00–1:30','1:30–2:00','2:00–2:30','2:30–3:00','3:00–3:30']

raw = json.load(open(r"C:\setag.mx\asignacion.json",encoding="utf-8"))
assign = {}
for k,v in raw.items():
    g,d,p = k.split("|"); assign[(g,d,p)] = v

# --- Verificación de traslapes de especialistas ---
def teacher_key(g, subj):
    n = 'baja' if grade(g)<=3 else 'alta'
    return {'Inglés':('Inglés',n),'Socioemocional':('Socio',n),'Educación Física':('EF','t'),
            'Desarrollo Comunitario':('DC','t'),'Computación':('Comp','t')}.get(subj)
occ = {}
for (g,d,p),s in assign.items():
    tk = teacher_key(g,s)
    if not tk: continue
    start = dict(blocks_for(turno(g),d))[p]
    occ.setdefault((tk,d,start),[]).append(g)
clash = {k:v for k,v in occ.items() if len(v)>1}
print("Traslapes de especialistas:", "NINGUNO ✅" if not clash else clash)
# Computación solo Mié/Jue
bad = [(g,d,p) for (g,d,p),s in assign.items() if s=='Computación' and d not in ('Mié','Jue')]
print("Computación fuera de Mié/Jue:", "NINGUNO ✅" if not bad else bad)

# --- Horas administrativas comunes por grado (2): ambos titulares libres a la vez ---
SPEC_SET = {'Inglés','Socioemocional','Educación Física','Desarrollo Comunitario','Computación'}
LBL = {480:'8:00-9:00',540:'9:00-10:00',630:'10:30-11:30',690:'11:30-12:30',
       750:'12:30-1:30',810:'1:30-2:30',870:'2:30-3:30'}
admin_slots = {}   # grado -> [(día,pos,start), ...]  (exactamente 2)
for gr in range(1,7):
    A, B = f"{gr}A", f"{gr}B"
    t = 1 if gr <= 3 else 2
    common = []
    for d in DAYS:
        for (p,start) in blocks_for(t,d):
            if assign[(A,d,p)] in SPEC_SET and assign[(B,d,p)] in SPEC_SET:
                common.append((d,p,start))
    order = {x:i for i,x in enumerate(DAYS)}
    common.sort(key=lambda c: (order[c[0]], c[2]))
    admin_slots[gr] = common[:2]

def admin_ticks_for(g):
    res = set()
    for (d,p,start) in admin_slots[grade(g)]:
        t0,t1 = START_TICKS[start]
        res.add((t0, d)); res.add((t1, d))   # ambas filas del bloque de 1h
    return res

# --- Orphan (30 min) Fábrica / Ortografía ---
ORPHAN = {'Lun':'Fábrica de lectura','Mar':'Ortografía','Mié':'Fábrica de lectura','Jue':'Ortografía','Vie':None}

def orphan_tick(g, day):
    if day=='Vie': return 11
    return 9 if turno(g)==1 else 11
def comida_tick(g, day):
    if day=='Vie': return None
    return 10 if turno(g)==1 else 12

# --- Construir grid de ticks por grupo/día ---
def build_grid(g):
    grid = {}  # (tick,day) -> (text, kind)
    for day in DAYS:
        for (pos,start) in blocks_for(turno(g),day):
            t0,t1 = START_TICKS[start]
            subj = assign[(g,day,pos)]
            grid[(t0,day)] = (subj,'clase'); grid[(t1,day)] = (subj,'clase_cont')
        grid[(4,day)] = ('RECREO','recreo')
        ct = comida_tick(g,day)
        if ct is not None: grid[(ct,day)] = ('COMIDA','comida')
        ot = orphan_tick(g,day)
        if ORPHAN[day]: grid[(ot,day)] = (ORPHAN[day]+' (30 min)','orphan')
        else: grid[(ot,day)] = ('Lectura libre (30 min)','orphan')
        if day=='Vie':
            for tk in (12,13,14): grid[(tk,day)] = ('SALIDA 2:00 PM','salida')
    return grid

# --- Colores por materia ---
COLORS = {
 'Lecto':'FCE4D6','Español':'DDEBF7','Matemáticas':'E2EFDA','Inglés':'FFF2CC','Socioemocional':'FBE5D6',
 'Desarrollo Comunitario':'E4DFEC','Educación Física':'D9E1F2','Computación':'EDEDED','Artes':'FCE4EC',
 'Formación':'FFF7E6','Conocimiento':'E2F0D9','Proyectos':'DEEBF7','Entidad':'F2DCDB','Historia':'F8CBAD',
 'Geografía':'D6E4F0','Fábrica de lectura':'FFF2CC','Ortografía':'FDE9D9'}
GREY='808080'; BREAKF='C9C9C9'; LUNCHF='FFD966'; SALIDAF='BFBFBF'

thin = Side(style='thin', color='B0B0B0')
border = Border(left=thin,right=thin,top=thin,bottom=thin)
adm = Side(style='medium', color='C00000')
admin_border = Border(left=adm,right=adm,top=adm,bottom=adm)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

wb = Workbook(); wb.remove(wb.active)

def style_cell(c, fill=None, bold=False, sz=10, color='000000'):
    c.alignment = center; c.border = border
    c.font = Font(bold=bold, size=sz, color=color)
    if fill: c.fill = PatternFill('solid', fgColor=fill)

for g in GROUPS:
    ws = wb.create_sheet(f"{g}")
    grid = build_grid(g)
    acells = admin_ticks_for(g)
    com = '1:00–1:30' if turno(g)==1 else '2:00–2:30'
    ws['A1'] = f"HORARIO {grade(g)}° \"{g[1]}\"  ·  Turno comida {com}  ·  Lun-Jue 8:00-3:30, Vie 8:00-2:00"
    ws.merge_cells('A1:F1'); style_cell(ws['A1'], '1F4E78', True, 12, 'FFFFFF')
    # encabezado
    ws['A2']='Hora'; style_cell(ws['A2'],'2E75B6',True,10,'FFFFFF')
    for j,day in enumerate(DAYS):
        c = ws.cell(row=2,column=2+j,value=day); style_cell(c,'2E75B6',True,10,'FFFFFF')
    # filas
    for i,lbl in enumerate(ROW_LABELS):
        r = 3+i
        c = ws.cell(row=r,column=1,value=lbl); style_cell(c,'D9D9D9',True,9)
        for j,day in enumerate(DAYS):
            txt,kind = grid.get((i,day),('',''))
            cell = ws.cell(row=r,column=2+j)
            if kind=='clase_cont':
                cell.value=''
            else:
                cell.value=txt
            if kind in ('clase','clase_cont'):
                style_cell(cell, COLORS.get(txt,'FFFFFF'), False, 9)
                if (i,day) in acells:
                    if kind=='clase':
                        cell.value = f"{txt}  ⚙\nHORA ADMIN."
                        cell.font = Font(bold=True, size=9, color='7F1D1D')
                    cell.border = admin_border
            elif kind=='recreo':
                style_cell(cell, BREAKF, True, 9, '404040')
            elif kind=='comida':
                style_cell(cell, LUNCHF, True, 9, '404040')
            elif kind=='orphan':
                style_cell(cell, COLORS.get(txt.split(' (')[0],'F2F2F2'), False, 8, '404040')
            elif kind=='salida':
                style_cell(cell, SALIDAF, True, 8, '404040')
            else:
                style_cell(cell, 'FFFFFF')
    # merge bloques de 1h (clase + clase_cont) y recreos/comidas/salidas verticales contiguas iguales
    for j,day in enumerate(DAYS):
        col = 2+j
        # merge pares de clase
        for start,(t0,t1) in START_TICKS.items():
            if (t0,day) in grid and grid[(t0,day)][1]=='clase':
                ws.merge_cells(start_row=3+t0,start_column=col,end_row=3+t1,end_column=col)
        # merge salida viernes (ticks 12-14)
        if day=='Vie':
            ws.merge_cells(start_row=3+12,start_column=col,end_row=3+14,end_column=col)
    ws.column_dimensions['A'].width = 13
    for j in range(len(DAYS)): ws.column_dimensions[get_column_letter(2+j)].width = 22
    ws.row_dimensions[1].height = 30
    for i in range(len(ROW_LABELS)): ws.row_dimensions[3+i].height = 22
    # leyenda hora administrativa
    lr = 3 + len(ROW_LABELS) + 1
    leg = ('⚙ HORA ADMIN. = el docente titular está libre (su grupo está con un especialista). '
           f'Coincide con el otro grupo de {grade(g)}° para trabajo administrativo conjunto (2 veces/semana). '
           'El recuadro rojo marca esos bloques.')
    ws.cell(row=lr, column=1, value=leg)
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=6)
    ws.cell(row=lr, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.cell(row=lr, column=1).font = Font(italic=True, size=8, color='7F1D1D')
    ws.row_dimensions[lr].height = 42

# --- Hoja RESUMEN ---
ws = wb.create_sheet("RESUMEN", 0)
ws['A1']='RESUMEN DE CUMPLIMIENTO DE CARGA HORARIA (horas/semana)'
ws.merge_cells('A1:N1'); style_cell(ws['A1'],'1F4E78',True,12,'FFFFFF')
# orden de materias
ALLSUBJ = ['Lecto','Español','Matemáticas','Inglés','Socioemocional','Desarrollo Comunitario',
           'Educación Física','Computación','Artes','Formación','Entidad','Conocimiento',
           'Proyectos','Fábrica de lectura','Ortografía','Historia','Geografía']
ws.cell(row=2,column=1,value='Materia'); style_cell(ws.cell(row=2,column=1),'2E75B6',True,9,'FFFFFF')
for j,g in enumerate(GROUPS):
    style_cell(ws.cell(row=2,column=2+j,value=g),'2E75B6',True,9,'FFFFFF')
style_cell(ws.cell(row=2,column=2+len(GROUPS),value='Total'),'2E75B6',True,9,'FFFFFF')

def scheduled_hours(g):
    h = {}
    for (gg,d,p),s in assign.items():
        if gg==g: h[s]=h.get(s,0)+1
    # orphan halves
    fab = sum(1 for d in DAYS if ORPHAN[d]=='Fábrica de lectura')*0.5
    ort = sum(1 for d in DAYS if ORPHAN[d]=='Ortografía')*0.5
    h['Fábrica de lectura']=fab; h['Ortografía']=ort
    return h

req = {}
for g in GROUPS:
    r = dict(curri[grade(g)]); r['Fábrica de lectura']=1; r['Ortografía']=1
    req[g]=r

all_ok=True
for i,subj in enumerate(ALLSUBJ):
    r = 3+i
    style_cell(ws.cell(row=r,column=1,value=subj),'D9D9D9',True,9); ws.cell(row=r,column=1).alignment=Alignment(horizontal='left',vertical='center')
    for j,g in enumerate(GROUPS):
        need = req[g].get(subj,0)
        got = scheduled_hours(g).get(subj,0)
        val = ('' if need==0 and got==0 else (f"{got:g}" if got==need else f"{got:g}/{need:g}✗"))
        if need!=got and not (need==0 and got==0): all_ok=False
        c=ws.cell(row=r,column=2+j,value=val); style_cell(c,'E2EFDA' if (need==got and need>0) else 'FFFFFF',False,9)
    tot = sum(scheduled_hours(g).get(subj,0) for g in GROUPS)
    style_cell(ws.cell(row=r,column=2+len(GROUPS),value=f"{tot:g}"),'FFF2CC',True,9)
# fila total por grupo
r = 3+len(ALLSUBJ)
style_cell(ws.cell(row=r,column=1,value='TOTAL'),'2E75B6',True,9,'FFFFFF')
for j,g in enumerate(GROUPS):
    tot = sum(scheduled_hours(g).values())
    style_cell(ws.cell(row=r,column=2+j,value=f"{tot:g}"),'C6E0B4',True,9)
style_cell(ws.cell(row=r,column=2+len(GROUPS),value=''),'2E75B6',True,9)
ws.column_dimensions['A'].width=22
for j in range(len(GROUPS)+1): ws.column_dimensions[get_column_letter(2+j)].width=7

# nota
r2=r+2
ws.cell(row=r2,column=1,value=('Notas: Fábrica de lectura y Ortografía se imparten en sesiones de 30 min (2/sem c/u). '
  'Computación solo Mié-Jue (1 docente). Inglés y Socioemocional: docente distinto para primaria baja (1-3) y alta (4-6). '
  'Ed. Física y Desarrollo Comunitario: 1 docente cada uno, sin traslapes. Matemáticas/Español/Lecto en bloques de 2h.'))
ws.merge_cells(start_row=r2,start_column=1,end_row=r2,end_column=2+len(GROUPS))
ws.cell(row=r2,column=1).alignment=Alignment(wrap_text=True,vertical='top'); ws.row_dimensions[r2].height=60

# --- Tabla de Horas Administrativas por grado (usa celdas combinadas, sin cambiar anchos) ---
ra = r2 + 2
style_cell(ws.cell(row=ra,column=1,value='⚙ HORAS ADMINISTRATIVAS COMUNES POR GRADO (titulares A y B libres a la vez)'),'C00000',True,10,'FFFFFF')
ws.merge_cells(start_row=ra,start_column=1,end_row=ra,end_column=2+len(GROUPS))
def adm_row(rr, c1, v1, c23, v2, c45, v3, fill1, fill2):
    style_cell(ws.cell(row=rr,column=1,value=v1),fill1,True,9, 'FFFFFF' if fill1=='7F1D1D' else '000000')
    ws.cell(row=rr,column=2,value=v2); ws.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=4)
    style_cell(ws.cell(row=rr,column=2),fill2,(fill2=='7F1D1D'),9,'FFFFFF' if fill2=='7F1D1D' else '000000')
    ws.cell(row=rr,column=5,value=v3); ws.merge_cells(start_row=rr,start_column=5,end_row=rr,end_column=7)
    style_cell(ws.cell(row=rr,column=5),fill2,(fill2=='7F1D1D'),9,'FFFFFF' if fill2=='7F1D1D' else '000000')
adm_row(ra+1, 1,'Grado', None,'Hora administrativa 1', None,'Hora administrativa 2', '7F1D1D','7F1D1D')
for k,gr in enumerate(range(1,7)):
    s = admin_slots[gr]
    v2 = f"{s[0][0]}  {LBL[s[0][2]]}" if len(s)>0 else '—'
    v3 = f"{s[1][0]}  {LBL[s[1][2]]}" if len(s)>1 else '—'
    adm_row(ra+2+k, 1,f"{gr}° (A y B)", None,v2, None,v3, 'F2DCDB','FBE5E5')

print("Horas administrativas por grado:")
for gr in range(1,7):
    print(f"  {gr}°:", ", ".join(f"{d} {LBL[s]}" for (d,p,s) in admin_slots[gr]))
print("Cumplimiento 100%:", "SÍ ✅" if all_ok else "NO ❌")
wb.save(r"C:\setag.mx\Horario_Escolar_2026.xlsx")
print("Excel guardado: C:\\setag.mx\\Horario_Escolar_2026.xlsx")
